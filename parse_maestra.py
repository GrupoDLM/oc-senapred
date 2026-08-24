#!/usr/bin/env python3
"""Convierte la maestra del CM Emergencias (EMERGENCIAS.xlsx) en un catálogo
compacto productos.json para el buscador del dashboard.
Uso: python3 parse_maestra.py EMERGENCIAS.xlsx oc-dashboard/data/productos.json
Reutilizable por el robot (GitHub Actions)."""
import openpyxl, os, json, sys

def cl(s): return ("" if s is None else str(s)).strip()
def num(s):
    s=cl(s)
    if s in ("", "None"): return None
    try: return int(round(float(s.replace(".","").replace(",","."))))
    except: return None

def parse(xlsx_path, out_path, actualizado=""):
    wb=openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws=wb["Datos"] if "Datos" in wb.sheetnames else wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True)
    hdr=list(next(it)); idx={h:i for i,h in enumerate(hdr)}
    def g(row,name):
        i=idx.get(name); return row[i] if (i is not None and i<len(row)) else None
    prods={}   # id_configurable -> product
    for row in it:
        if row is None or all(v is None for v in row): continue
        cid=cl(g(row,"ID CONFIGURABLE"))
        nombre=cl(g(row,"PRODUCTO CONFIGURABLE")) or cl(g(row,"PRODUCTO"))
        reg=cl(g(row,"REGIÓN")); precio=num(g(row,"PRECIO EN TIENDA"))
        if not cid or not nombre or not reg or precio is None: continue
        p=prods.get(cid)
        if not p:
            p={"id":cid,"nombre":nombre,"tipo":cl(g(row,"TIPO PRODUCTO")),
               "cat":cl(g(row,"CATEGORÍA 2")),"modelo":cl(g(row,"MODELO")),
               "medida":cl(g(row,"MEDIDA")),"precios":{}}
            prods[cid]=p
        prov=cl(g(row,"NOMBRE PROVEEDOR"))
        cur=p["precios"].get(reg)
        # nos quedamos con el menor precio por región, guardando el proveedor
        if cur is None or precio < cur["p"]:
            p["precios"][reg]={"p":precio,"prov":prov}
    items=sorted(prods.values(), key=lambda x:(x["tipo"], x["nombre"]))
    tipos=sorted(set(i["tipo"] for i in items if i["tipo"]))
    regiones=sorted({r for i in items for r in i["precios"].keys()})
    out={"fuente":"Convenio Marco Emergencias y Prevención (2239-8-LR24)",
         "actualizado":actualizado,"tipos":tipos,"regiones":regiones,"items":items}
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    json.dump(out, open(out_path,"w",encoding="utf-8"), ensure_ascii=False, separators=(",",":"))
    return out, os.path.getsize(out_path)

if __name__=="__main__":
    xlsx=sys.argv[1] if len(sys.argv)>1 else "EMERGENCIAS.xlsx"
    out=sys.argv[2] if len(sys.argv)>2 else "oc-dashboard/data/productos.json"
    o,sz=parse(xlsx, out, actualizado=(sys.argv[3] if len(sys.argv)>3 else ""))
    print("Productos:", len(o["items"]), "| Tipos:", len(o["tipos"]),
          "| Regiones:", len(o["regiones"]), "| %.1f KB"%(sz/1024))
