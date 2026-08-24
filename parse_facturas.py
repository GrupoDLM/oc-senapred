#!/usr/bin/env python3
"""Lee un Excel con columnas 'OC' y 'N° Factura' y arma facturas.json:
   {actualizado, map:{ "<codigoOC>": ["<fact>", ...] }}.
   Una OC puede tener varias facturas (entregas parciales, refacturación, etc.).
   Reutilizable por el robot / futura conexión SII.
Uso: python3 parse_facturas.py ARCHIVO.xlsx oc-dashboard/data/facturas.json [YYYY-MM-DD]"""
import openpyxl, os, json, sys, collections

NA={"","#N/A","N/A","#N/D","#REF!","0","SIN FACTURA","-"}
def cl(v): return ("" if v is None else str(v)).strip()

def parse(xlsx, out, actualizado=""):
    wb=openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    # elige la hoja que tenga columna "N° Factura"
    target=None
    for sh in wb.sheetnames:
        ws=wb[sh]; hdr=[cl(c) for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
        if any(h.lower().replace("º","°").startswith("n° factura") or "factura" in h.lower() for h in hdr):
            target=(sh,hdr); break
    if not target: raise SystemExit("No encontré columna de Factura en el Excel")
    sh,hdr=target; ws=wb[sh]; idx={h:i for i,h in enumerate(hdr)}
    iOC=idx.get("OC");
    iF=next((i for h,i in idx.items() if "factura" in h.lower()),None)
    it=ws.iter_rows(values_only=True); next(it)
    mp=collections.OrderedDict()
    for r in it:
        if r is None or all(v is None for v in r): continue
        oc=cl(r[iOC]); f=cl(r[iF]) if iF is not None else ""
        if not oc: continue
        if f.upper() in NA: continue
        mp.setdefault(oc,[])
        if f not in mp[oc]: mp[oc].append(f)
    out_obj={"actualizado":actualizado,"map":mp}
    os.makedirs(os.path.dirname(out), exist_ok=True)
    json.dump(out_obj, open(out,"w",encoding="utf-8"), ensure_ascii=False, separators=(",",":"))
    return out_obj, os.path.getsize(out)

if __name__=="__main__":
    xlsx=sys.argv[1] if len(sys.argv)>1 else "FACT.xlsx"
    out=sys.argv[2] if len(sys.argv)>2 else "oc-dashboard/data/facturas.json"
    o,sz=parse(xlsx,out,actualizado=(sys.argv[3] if len(sys.argv)>3 else ""))
    print("OCs con factura:",len(o["map"]),"| %.1f KB"%(sz/1024))
